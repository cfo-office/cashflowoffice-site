#!/usr/bin/env python3
"""Automated structural audit for Cash Flow Office Excel workbooks."""

from __future__ import annotations

import json
import re
import sys
import zipfile
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from xml.etree import ElementTree

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[2]
SOURCE_DIR = ROOT / "digital-products" / "excel-source"
REPORT_DIR = ROOT / "digital-products" / "qc-reports"
JSON_REPORT = REPORT_DIR / "automated-workbook-audit.json"
MD_REPORT = REPORT_DIR / "automated-workbook-audit.md"

ERROR_STRINGS = ["#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A", "#NUM!", "#NULL!"]
EXTERNAL_REF_RE = re.compile(r"(?:'?\[[^\]]+\][^']*'?!)|(?:https?://)|(?:file://)", re.IGNORECASE)
BROKEN_REF_RE = re.compile(r"#REF!", re.IGNORECASE)
SHEET_REF_RE = re.compile(r"(?:'([^']+)'|([A-Za-z0-9_ .-]+))!")


def now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def cell_ref(cell: Cell) -> str:
    return f"{cell.parent.title}!{cell.coordinate}"


def is_formula(cell: Cell) -> bool:
    return isinstance(cell.value, str) and cell.value.startswith("=")


def is_error_string(value: Any) -> bool:
    return isinstance(value, str) and value in ERROR_STRINGS


def style_locked(cell: Cell) -> bool:
    protection = getattr(cell, "protection", None)
    if protection is None or protection.locked is None:
        return True
    return bool(protection.locked)


def safe_len(value: Any) -> int:
    try:
        return len(value)
    except TypeError:
        return 0


def sample(items: list[Any], limit: int = 50) -> list[Any]:
    return items[:limit]


def get_calculation_mode(workbook: Any) -> str | None:
    calculation = getattr(workbook, "calculation", None)
    if calculation is not None:
        calc_mode = getattr(calculation, "calcMode", None)
        if calc_mode:
            return str(calc_mode)

    properties = getattr(workbook, "calculation_properties", None)
    if properties is not None:
        calc_mode = getattr(properties, "calcMode", None)
        if calc_mode:
            return str(calc_mode)

    return None


def get_defined_names(workbook: Any) -> list[dict[str, Any]]:
    defined_names: list[dict[str, Any]] = []
    names = getattr(workbook, "defined_names", None)

    if names is None:
        return defined_names

    try:
        iterable = names.items()
    except AttributeError:
        iterable = []

    for name, definition in iterable:
        destinations = []
        try:
            destinations = [
                {"sheet": sheet, "range": cell_range}
                for sheet, cell_range in definition.destinations
            ]
        except Exception:
            destinations = []

        defined_names.append({
            "name": name,
            "value": getattr(definition, "attr_text", None),
            "scope": getattr(definition, "localSheetId", None),
            "destinations": destinations
        })

    return defined_names


def detect_external_links_from_zip(path: Path) -> list[str]:
    links: set[str] = set()

    try:
        with zipfile.ZipFile(path) as archive:
            for name in archive.namelist():
                lowered = name.lower()
                if lowered.startswith("xl/externallinks/"):
                    links.add(name)

                if lowered.endswith(".rels"):
                    try:
                        root = ElementTree.fromstring(archive.read(name))
                    except Exception:
                        continue

                    for relationship in root:
                        target = relationship.attrib.get("Target", "")
                        rel_type = relationship.attrib.get("Type", "")
                        target_mode = relationship.attrib.get("TargetMode", "")
                        if (
                            "externalLink" in rel_type
                            or target_mode.lower() == "external"
                            or target.startswith(("http://", "https://", "file://"))
                        ):
                            links.add(target or name)
    except Exception:
        return sorted(links)

    return sorted(links)


def iter_used_cells(sheet: Any) -> list[Cell]:
    return [
        cell
        for row in sheet.iter_rows(
            min_row=1,
            max_row=sheet.max_row,
            min_col=1,
            max_col=sheet.max_column,
        )
        for cell in row
    ]


def detect_repeated_formula_patterns(formula_cells: list[Cell]) -> list[dict[str, Any]]:
    by_formula: dict[str, list[str]] = defaultdict(list)
    for cell in formula_cells:
        by_formula[str(cell.value)].append(cell.coordinate)

    repeated = [
        {
            "formula": formula,
            "count": len(locations),
            "sample_cells": sample(locations, 12)
        }
        for formula, locations in by_formula.items()
        if len(locations) >= 3
    ]

    repeated.sort(key=lambda item: item["count"], reverse=True)
    return repeated[:20]


def detect_row_formula_patterns(formula_cells: list[Cell]) -> list[dict[str, Any]]:
    by_row: dict[int, list[Cell]] = defaultdict(list)
    by_column: dict[int, list[Cell]] = defaultdict(list)

    for cell in formula_cells:
        by_row[cell.row].append(cell)
        by_column[cell.column].append(cell)

    row_patterns = [
        {
            "direction": "across_row",
            "row": row,
            "count": len(cells),
            "columns": [get_column_letter(cell.column) for cell in cells[:20]]
        }
        for row, cells in by_row.items()
        if len(cells) >= 3
    ]

    column_patterns = [
        {
            "direction": "down_column",
            "column": get_column_letter(column),
            "count": len(cells),
            "rows": [cell.row for cell in cells[:20]]
        }
        for column, cells in by_column.items()
        if len(cells) >= 3
    ]

    patterns = row_patterns + column_patterns
    patterns.sort(key=lambda item: item["count"], reverse=True)
    return patterns[:25]


def suspicious_broken_sheet_refs(formula: str, sheet_names: set[str]) -> list[str]:
    refs: list[str] = []
    for quoted, bare in SHEET_REF_RE.findall(formula):
        sheet_name = quoted or bare
        if "[" in sheet_name:
            continue
        if sheet_name and sheet_name not in sheet_names and sheet_name.upper() not in {"TRUE", "FALSE"}:
            refs.append(sheet_name)
    return refs


def audit_sheet(sheet: Any, sheet_names: set[str]) -> dict[str, Any]:
    used_cells = iter_used_cells(sheet)
    filled_cells = [cell for cell in used_cells if cell.value is not None]
    formula_cells = [cell for cell in filled_cells if is_formula(cell)]
    unlocked_cells = [cell for cell in used_cells if not style_locked(cell)]
    locked_cells = [cell for cell in used_cells if style_locked(cell)]
    formula_unlocked = [cell for cell in formula_cells if not style_locked(cell)]
    locked_blank_cells = [cell for cell in used_cells if cell.value is None and style_locked(cell)]
    input_looking_locked = [
        cell
        for cell in filled_cells
        if (
            style_locked(cell)
            and not is_formula(cell)
            and isinstance(cell.value, (int, float, str))
            and not is_error_string(cell.value)
        )
    ]

    visible_errors = [
        {"cell": cell_ref(cell), "value": cell.value}
        for cell in filled_cells
        if is_error_string(cell.value)
    ]

    external_formula_refs = [
        {"cell": cell_ref(cell), "formula": cell.value}
        for cell in formula_cells
        if EXTERNAL_REF_RE.search(str(cell.value))
    ]

    ref_formula_errors = [
        {"cell": cell_ref(cell), "formula": cell.value}
        for cell in formula_cells
        if BROKEN_REF_RE.search(str(cell.value))
    ]

    broken_sheet_refs = []
    for cell in formula_cells:
        missing_sheets = suspicious_broken_sheet_refs(str(cell.value), sheet_names)
        if missing_sheets:
            broken_sheet_refs.append({
                "cell": cell_ref(cell),
                "formula": cell.value,
                "missing_sheet_references": sorted(set(missing_sheets))
            })

    data_validation_count = safe_len(getattr(sheet.data_validations, "dataValidation", []))
    conditional_formatting_count = safe_len(sheet.conditional_formatting)
    merged_ranges_count = safe_len(sheet.merged_cells.ranges)
    sheet_protected = bool(getattr(sheet.protection, "sheet", False))

    issues = {
        "critical": [],
        "important": [],
        "polish": []
    }

    if visible_errors:
        issues["critical"].append(f"Visible Excel error strings found: {len(visible_errors)}")
    if ref_formula_errors:
        issues["critical"].append(f"Formula text contains #REF!: {len(ref_formula_errors)}")
    if broken_sheet_refs:
        issues["critical"].append(f"Suspicious missing sheet references found: {len(broken_sheet_refs)}")
    if not sheet_protected:
        issues["important"].append("Sheet is unprotected.")
    if formula_unlocked:
        issues["important"].append(f"Formula cells are unlocked: {len(formula_unlocked)}")
    if data_validation_count == 0 and len(filled_cells) >= 10:
        issues["polish"].append("No data validations detected; confirm dropdown/input controls are not expected.")
    if len(formula_cells) == 0 and len(filled_cells) >= 10:
        issues["polish"].append("No formulas detected; confirm this sheet is intended to be input/reference only.")
    if input_looking_locked:
        issues["polish"].append("Some non-formula filled cells are locked; confirm intended input cells remain editable.")

    return {
        "sheet_name": sheet.title,
        "max_row": sheet.max_row,
        "max_column": sheet.max_column,
        "visibility": sheet.sheet_state,
        "merged_ranges_count": merged_ranges_count,
        "formula_cell_count": len(formula_cells),
        "visible_formula_error_strings": visible_errors,
        "data_validation_count": data_validation_count,
        "conditional_formatting_rule_count": conditional_formatting_count,
        "sheet_protection_enabled": sheet_protected,
        "freeze_pane": str(sheet.freeze_panes) if sheet.freeze_panes else None,
        "approximate_locked_cell_count": len(locked_cells),
        "approximate_unlocked_cell_count": len(unlocked_cells),
        "approximate_filled_cell_count": len(filled_cells),
        "unlocked_cells": sample([cell_ref(cell) for cell in unlocked_cells], 150),
        "formula_cells_not_locked": sample([cell_ref(cell) for cell in formula_unlocked], 100),
        "locked_blank_or_input_looking_cells": sample([cell_ref(cell) for cell in locked_blank_cells + input_looking_locked], 150),
        "formulas_containing_external_references": external_formula_refs,
        "formulas_containing_ref_errors": ref_formula_errors,
        "suspicious_broken_sheet_references": broken_sheet_refs,
        "repeated_exact_formula_patterns": detect_repeated_formula_patterns(formula_cells),
        "formula_distribution_patterns": detect_row_formula_patterns(formula_cells),
        "issues": issues
    }


def audit_workbook(path: Path) -> dict[str, Any]:
    result: dict[str, Any] = {
        "filename": path.name,
        "path": str(path.relative_to(ROOT)),
        "opens_successfully": False,
        "sheet_names": [],
        "visible_sheet_count": 0,
        "hidden_sheet_count": 0,
        "hidden_sheet_names": [],
        "external_workbook_links": detect_external_links_from_zip(path),
        "defined_named_ranges": [],
        "workbook_calculation_mode": None,
        "sheets": [],
        "issues": {
            "critical": [],
            "important": [],
            "polish": []
        }
    }

    try:
        workbook = load_workbook(path, data_only=False, keep_links=True)
    except Exception as exc:
        result["error"] = f"{type(exc).__name__}: {exc}"
        result["issues"]["critical"].append("Workbook could not be opened by openpyxl.")
        return result

    result["opens_successfully"] = True
    result["sheet_names"] = workbook.sheetnames
    result["visible_sheet_count"] = sum(1 for sheet in workbook.worksheets if sheet.sheet_state == "visible")
    result["hidden_sheet_count"] = sum(1 for sheet in workbook.worksheets if sheet.sheet_state != "visible")
    result["hidden_sheet_names"] = [sheet.title for sheet in workbook.worksheets if sheet.sheet_state != "visible"]
    result["defined_named_ranges"] = get_defined_names(workbook)
    result["workbook_calculation_mode"] = get_calculation_mode(workbook)

    sheet_names = set(workbook.sheetnames)
    result["sheets"] = [audit_sheet(sheet, sheet_names) for sheet in workbook.worksheets]

    if result["external_workbook_links"]:
        result["issues"]["critical"].append("External workbook links detected.")
    if result["hidden_sheet_count"]:
        result["issues"]["important"].append(f"Hidden sheets detected: {', '.join(result['hidden_sheet_names'])}")
    if result["workbook_calculation_mode"] not in (None, "auto"):
        result["issues"]["important"].append(f"Workbook calculation mode is {result['workbook_calculation_mode']}.")
    if result["defined_named_ranges"]:
        result["issues"]["polish"].append(f"Defined names/ranges present: {len(result['defined_named_ranges'])}.")

    for sheet in result["sheets"]:
        for severity in ("critical", "important", "polish"):
            for issue in sheet["issues"][severity]:
                result["issues"][severity].append(f"{sheet['sheet_name']}: {issue}")

    return result


def summarize(results: list[dict[str, Any]]) -> dict[str, Any]:
    totals = {
        "workbooks_found": len(results),
        "workbooks_opened": sum(1 for workbook in results if workbook["opens_successfully"]),
        "critical_issue_count": 0,
        "important_issue_count": 0,
        "polish_item_count": 0,
        "formula_cell_count": 0,
        "visible_error_count": 0,
        "external_formula_reference_count": 0,
        "unprotected_sheet_count": 0,
        "formula_unlocked_count": 0
    }

    for workbook in results:
        totals["critical_issue_count"] += len(workbook["issues"]["critical"])
        totals["important_issue_count"] += len(workbook["issues"]["important"])
        totals["polish_item_count"] += len(workbook["issues"]["polish"])

        for sheet in workbook.get("sheets", []):
            totals["formula_cell_count"] += sheet["formula_cell_count"]
            totals["visible_error_count"] += len(sheet["visible_formula_error_strings"])
            totals["external_formula_reference_count"] += len(sheet["formulas_containing_external_references"])
            totals["formula_unlocked_count"] += len(sheet["formula_cells_not_locked"])
            if not sheet["sheet_protection_enabled"]:
                totals["unprotected_sheet_count"] += 1

    return totals


def issue_lines(results: list[dict[str, Any]], severity: str) -> list[str]:
    lines: list[str] = []
    for workbook in results:
        for issue in workbook["issues"][severity]:
            lines.append(f"- `{workbook['filename']}`: {issue}")
    return lines


def render_markdown(report: dict[str, Any]) -> str:
    results = report["workbooks"]
    summary = report["summary"]

    lines = [
        "# Automated Workbook Audit",
        "",
        f"Generated: {report['generated_at']}",
        "",
        "This is an automated structural audit. It is not a replacement for manual Excel scenario testing, workbook usability review, or customer workflow validation.",
        "",
        "## Executive Summary",
        "",
        f"- Workbooks found: {summary['workbooks_found']}",
        f"- Workbooks opened successfully: {summary['workbooks_opened']}",
        f"- Total formula cells scanned: {summary['formula_cell_count']}",
        f"- Visible formula/error strings found: {summary['visible_error_count']}",
        f"- External formula references found: {summary['external_formula_reference_count']}",
        f"- Unprotected sheets found: {summary['unprotected_sheet_count']}",
        f"- Formula cells not locked: {summary['formula_unlocked_count']}",
        f"- Critical issues: {summary['critical_issue_count']}",
        f"- Important issues: {summary['important_issue_count']}",
        f"- Polish/manual review items: {summary['polish_item_count']}",
        "",
        "## Workbook-by-Workbook Status",
        ""
    ]

    for workbook in results:
        status = "Opened" if workbook["opens_successfully"] else "Failed to open"
        lines.extend([
            f"### {workbook['filename']}",
            "",
            f"- Status: {status}",
            f"- Sheets: {', '.join(workbook['sheet_names']) if workbook['sheet_names'] else 'None detected'}",
            f"- Visible sheets: {workbook['visible_sheet_count']}",
            f"- Hidden sheets: {workbook['hidden_sheet_count']}",
            f"- Hidden sheet names: {', '.join(workbook['hidden_sheet_names']) if workbook['hidden_sheet_names'] else 'None'}",
            f"- External workbook links: {', '.join(workbook['external_workbook_links']) if workbook['external_workbook_links'] else 'None detected'}",
            f"- Defined/named ranges: {len(workbook['defined_named_ranges'])}",
            f"- Calculation mode: {workbook['workbook_calculation_mode'] or 'Not detected'}",
            ""
        ])

        if workbook["sheets"]:
            lines.append("| Sheet | Size | State | Formulas | Errors | Validations | Conditional Formats | Protected | Freeze Pane | Locked / Unlocked / Filled |")
            lines.append("| --- | --- | --- | ---: | ---: | ---: | ---: | --- | --- | --- |")
            for sheet in workbook["sheets"]:
                lines.append(
                    f"| {sheet['sheet_name']} | {sheet['max_row']}x{sheet['max_column']} | {sheet['visibility']} | "
                    f"{sheet['formula_cell_count']} | {len(sheet['visible_formula_error_strings'])} | "
                    f"{sheet['data_validation_count']} | {sheet['conditional_formatting_rule_count']} | "
                    f"{'Yes' if sheet['sheet_protection_enabled'] else 'No'} | {sheet['freeze_pane'] or 'None'} | "
                    f"{sheet['approximate_locked_cell_count']} / {sheet['approximate_unlocked_cell_count']} / {sheet['approximate_filled_cell_count']} |"
                )
            lines.append("")

    for heading, severity in [
        ("Critical Issues", "critical"),
        ("Important Issues", "important"),
        ("Polish / Manual Review Items", "polish")
    ]:
        lines.extend([f"## {heading}", ""])
        items = issue_lines(results, severity)
        if items:
            lines.extend(items)
        else:
            lines.append("- None detected by automated scan.")
        lines.append("")

    lines.extend([
        "## Recommended Next Actions",
        "",
        "- Manually open every workbook in Excel and run scenario tests using realistic builder/contractor data.",
        "- Confirm all intended input cells are editable and all intended formula cells are protected.",
        "- Confirm unprotected sheets are intentional or protect them before release.",
        "- Confirm sheets with no data validations do not require dropdown controls.",
        "- Confirm formula counts make sense for each workbook and that low-formula sheets are intended to be input/reference sheets.",
        "- Re-run this audit after any workbook edits and before rebuilding the release ZIP.",
        ""
    ])

    return "\n".join(lines)


def main() -> int:
    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    workbook_paths = sorted(SOURCE_DIR.glob("*.xlsx"))

    results = [audit_workbook(path) for path in workbook_paths]
    report = {
        "generated_at": now_iso(),
        "source_dir": str(SOURCE_DIR.relative_to(ROOT)),
        "audit_note": "Automated structural audit only; manual Excel scenario testing is still required.",
        "summary": summarize(results),
        "workbooks": results
    }

    JSON_REPORT.write_text(json.dumps(report, indent=2), encoding="utf-8")
    MD_REPORT.write_text(render_markdown(report), encoding="utf-8")

    print(f"Wrote {JSON_REPORT.relative_to(ROOT)}")
    print(f"Wrote {MD_REPORT.relative_to(ROOT)}")
    print(json.dumps(report["summary"], indent=2))
    return 0


if __name__ == "__main__":
    sys.exit(main())
